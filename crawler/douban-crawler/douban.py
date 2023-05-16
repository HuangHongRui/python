import time
import urllib
import requests
import numpy as np
from bs4 import BeautifulSoup
from openpyxl import Workbook

headers = [
  { 'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36' },
  { 'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.1 Safari/605.1.15' },
  { 'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36 Edg/113.0.1774.42' }
]

def book_spider(tag):
  page_num = 0
  try_times = 0
  book_list = []

  while (1):
    #url='http://www.douban.com/tag/小说/book?start=0' # For Test
    url = 'https://www.douban.com/tag/' + tag + '/book?start=' + str(page_num * 15)

    # 程序在执行到这一行时暂停一段随机的时间，范围在 0 到 5 秒之间。这样可以模拟爬取网页时的随机间隔，以避免对目标网站造成过大的请求压力。
    time.sleep(np.random.rand() * 5)

    try:
      response = requests.get(url, headers=headers[page_num % len(headers)])
      content = response.text
    except:
      continue

    soup = BeautifulSoup(content)

    # 找到class:mod book-list
    list_soup = soup.find('div', {'class': 'mod book-list'})
    try_times += 1

    if list_soup == None and try_times < 200:
      continue
    elif list_soup == None or len(list_soup) <= 1:
      break

    if page_num >= 1:
      break

    for book_info in list_soup.findAll('dd'):
      title = book_info.find('a', {'class': 'title'}).string.strip()
      desc = book_info.find('div', {'class': 'desc'}).string.strip()
      desc_list = desc.split('/')
      book_url = book_info.find('a', {'class': 'title'}).get('href')

      try:
        author_info = '作者/译者：' + '/'.join(desc_list[0: -3])
      except:
        author_info = '作者/译者：暂无'

      try:
        pub_info = '出版信息：' + '/'.join(desc_list[-3:])
      except:
        pub_info = '出版信息：暂无'

      try:
          rating = book_info.find('span', {'class':'rating_nums'}).string.strip()
      except:
          rating='0.0'

      try:
          people_num = get_people_num(book_url)
          people_num = people_num.strip('人评价')
          people_num  = '0'
      except:
          people_num ='0'

      book_list.append([title, rating, people_num, author_info, pub_info])
      try_times=0 #set 0 when got valid information
    
    page_num += 1

    print('从页面下载信息 %d' % page_num)

  return book_list

# 获取每本书的评价人数
def get_people_num(url):
    #url='http://book.douban.com/subject/6082808/?from=tag_all' # For Test
    try:
        response = requests.get(url, headers=headers[np.random.randint(0, len(headers))])
        content = response.text
    except:
        print('请求评价人数失败', url)

    soup = BeautifulSoup(content)
    people_num = soup.find('div',{'class':'rating_sum'}).findAll('span')[1].string.strip()
    return people_num

# 执行爬虫
# book_list 是一个包含书籍信息的列表，通过以平分作为关键字进行排序，按照从高到低的顺序进行排序。
def do_spider(book_tag_lists):
    book_lists = []
    for book_tag in book_tag_lists:
        book_list = book_spider(book_tag)
        book_list = sorted(book_list, key=lambda x:x[1], reverse=True)
        book_lists.append(book_list)
    return book_lists

# 存储到excel中
def print_book_lists_excel(book_lists, book_tag_lists):
    wb = Workbook(write_only = True)
    ws = []
    for i in range(len(book_tag_lists)):
        ws.append(wb.create_sheet(title = book_tag_lists[i]))

    for i in range(len(book_tag_lists)): 
        ws[i].append(['序号', '书名', '评分', '评价人数', '作者', '出版社'])
        count = 1
        for bl in book_lists[i]:
            ws[i].append([count, bl[0], float(bl[1]), int(bl[2]), bl[3], bl[4]])
            count += 1

    save_path = 'book_list'
    for i in range(len(book_tag_lists)):
        save_path += ('-' + book_tag_lists[i])

    save_path += '.xlsx'
    wb.save(save_path)
    
if __name__ == '__main__':
  book_tag_lists = ['名著']
  # book_tag_lists = ['个人管理', '时间管理', '投资', '文化', '宗教']
  book_lists = do_spider(book_tag_lists)
  print_book_lists_excel(book_lists, book_tag_lists)